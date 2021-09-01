import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import os
from selenium.webdriver.chrome.options import Options
import sys
from RPA.Browser.Selenium import Selenium


CURRENT_DIR = os.path.abspath(os.getcwd())
OUTPUT_DIR_NAME = 'output'
XLSX_FILE_NAME = 'result.xlsx'


class SeleniumDriver:
    def __init__(self):
        self.driver = None
        self.options = Options()
        self.options.add_experimental_option(
            'prefs', {
                "download.default_directory": os.path.join(CURRENT_DIR, 'output/'),
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True
            }
        )
        self.driver = Selenium()

    def scroll_to_element(self, xpath):
        locator = f'xpath:{xpath}'
        self.driver.scroll_element_into_view(locator)

    def visit_url(self, url):
        self.driver.open_available_browser(url=url)

    def check_is_element_present_by_xpath(self, xpath, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )

    def wait_until_element_presents(self, locator, time_wait):
        self.driver.wait_until_element_is_visible(locator, time_wait)

    def wait_until_disapear(self, class_name, time_wait):
        locator = f'css:.{class_name}'
        self.driver.wait_until_element_is_not_visible(locator, time_wait)

    def check_is_element_present_by_xpath_v2(self, xpath, time_to_wait=10):
        locator = f'xpath:{xpath}'
        self.driver.wait_until_element_is_visible(locator)

    def check_is_element_is_not_present_by_xpath(self, xpath, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.invisibility_of_element_located((By.XPATH, xpath))
        )

    def is_loading(self, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.invisibility_of_element_located((By.CLASS_NAME, 'loading'))
        )

    def search_for(self, term):
        input_field = "css:input"
        self.driver.input_text(input_field, term)
        self.driver.press_keys(input_field, "ENTER")

    def click(self, element):
        self.driver.click_element(element)

    def find_element_if_exists_by_xpath(self, xpath):
        selector = f'xpath:{xpath}'
        return self.driver.find_element(selector)

    def fetch_elements_if_exists_by_xpath(self, xpath):
        locator = f'xpath:{xpath}'
        return self.driver.find_elements(locator)

    def quit(self):
        self.driver.close_all_browsers()


class ExcelHandler:
    def __init__(self):
        self.wb = Workbook()
        self.destination_file_name = f'{OUTPUT_DIR_NAME}/{XLSX_FILE_NAME}'

    @staticmethod
    def check_or_create_folder():
        if not os.path.exists(OUTPUT_DIR_NAME):
            os.makedirs(OUTPUT_DIR_NAME)

    def initialize(self):
        self.check_or_create_folder()
        self.ws1 = self.wb.active
        self.ws1.title = 'Agencies'
        self.ws1.append([
            'Agency name',
            'Spend Amounts'
        ])

    def save(self):
        self.wb.save(filename=self.destination_file_name)

    def write_to_file(self, dict_items, sheet):
        for key, value in dict_items:
            sheet.append([key, value])

    def read_from_file(self):
        pass

    def create_sheet(self, title):
        return self.wb.create_sheet(title=title)


class ItDashboardScraper:
    def __init__(self, agency_name):
        self.links_to_download_pdf = []
        self.agencies = {}
        self.url = 'https://itdashboard.gov/'
        self.dive_in_xpath = '//*[@id="node-23"]/div/div/div/div/div/div/div/a'
        self.agency_item_xpath = '//*[@id="agency-tiles-widget"]//a/img'
        self.agency_items_xpath = '//*[@id="agency-tiles-widget"]/div/div/div/div/div/div/div[1]/a'
        self.table_element_xpath = '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[1]'
        self.select_element_xpath = '//*[@id="investments-table-object_length"]/label/select'
        self.all_options_xpath = '//*[@id="investments-table-object_length"]/label/select/option[4]'
        self.agency_name = agency_name
        self.agency_obj = None
        self.downloaded_pdf_file_names = []
        self.xlsx_handler = ExcelHandler()
        self.selenium_driver = SeleniumDriver()

    def parse_agencies(self):
        dive_in_button = self.selenium_driver.find_element_if_exists_by_xpath(self.dive_in_xpath)
        self.selenium_driver.click(dive_in_button)

        self.selenium_driver.check_is_element_present_by_xpath_v2(self.agency_item_xpath)
        agencies_elements = self.selenium_driver.fetch_elements_if_exists_by_xpath(self.agency_items_xpath)
        for item in agencies_elements:
            name = item.find_element_by_xpath('./span[1]').text
            value = item.find_element_by_xpath('./span[2]').text
            self.agencies[name] = value
            if name == self.agency_name:
                self.agency_obj = item


    def parse_table(self):
        # if self.selenium_driver.wait_until_disapear('loading'):
        self.selenium_driver.wait_until_disapear('loading', 20)
        table_body = self.selenium_driver.find_element_if_exists_by_xpath('//*[@id="investments-table-object"]/tbody')
        for row in table_body.find_elements_by_xpath('./tr'):
            _row = []
            for index, col in enumerate(row.find_elements_by_xpath('./td')):
                _row.append(col.text)
                link_text = col.find_elements_by_link_text(col.text)
                if link_text:
                    self.links_to_download_pdf.append(link_text[0].get_attribute('href'))
            self.ws2.append(_row)

    def check_agency(self):
        self.selenium_driver.click(self.agency_obj)
        self.selenium_driver.scroll_to_element('//*[@id="block-itdb-custom--5"]/div/div/div/div[2]/div/div[2]/h4')
        self.selenium_driver.wait_until_element_presents('css:select', 15)
        select_element = self.selenium_driver.find_element_if_exists_by_xpath('//*[@id="investments-table-object_length"]/label/select')
        self.selenium_driver.click(select_element)
        show_all_element = self.selenium_driver.find_element_if_exists_by_xpath('//*[@id="investments-table-object_length"]/label/select/option[4]')
        show_all_element.click()
        self.selenium_driver.click(show_all_element)
        self.parse_table()

    def download_pdf_files_from_links(self):
        for item in self.links_to_download_pdf:
            self.selenium_driver.visit_url(item)
            self.selenium_driver.wait_until_element_presents(f'xpath://*[@id="business-case-pdf"]/a', 10)
            element = self.selenium_driver.find_element_if_exists_by_xpath('//*[@id="business-case-pdf"]/a')
            self.selenium_driver.click(element)
            time.sleep(10)

    def execute(self):
        try:
            self.xlsx_handler.initialize()
            self.ws2 = self.xlsx_handler.create_sheet('individual investments')
            self.selenium_driver.visit_url(self.url)
            self.parse_agencies()
            self.xlsx_handler.write_to_file(self.agencies.items(), self.xlsx_handler.ws1)
            self.check_agency()
            self.download_pdf_files_from_links()
            self.xlsx_handler.save()
        finally:
            self.selenium_driver.quit()


if __name__ == '__main__':
    agency_name = 'National Science Foundation'
    web_page_data_extracter = ItDashboardScraper(agency_name=agency_name)
    web_page_data_extracter.execute()

