import time
from io import BytesIO

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import os
from selenium.webdriver.chrome.options import Options
import sys

current_dir = os.path.abspath(os.getcwd())


class SeleniumDriver:
    def __init__(self):
        self.driver = None
        self.options = Options()

    def set_options(self):
        self.options.add_experimental_option(
            'prefs', {
                "download.default_directory": os.path.join(current_dir, 'output/'),
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True
            }
        )

    def initialize_driver(self):
        self.set_options()
        self.driver = webdriver.Chrome(options=self.options)

    def check_is_element_present_by_xpath(self, xpath, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )

    def check_is_element_is_not_present_by_xpath(self, xpath, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.invisibility_of_element_located((By.XPATH, xpath))
        )

    def is_loading(self, time_to_wait=10):
        element = WebDriverWait(self.driver, time_to_wait).until(
            EC.invisibility_of_element_located((By.CLASS_NAME, 'loading'))
        )

    def get_element_if_exists_by_xpath(self, xpath):
        self.check_is_element_present_by_xpath(xpath)
        return self.driver.find_element_by_xpath(xpath)

    def fetch_elements_if_exists_by_xpath(self, xpath):
        return self.driver.find_elements_by_xpath(xpath)


class ExcelHandler:
    def __init__(self):
        self.wb = Workbook()
        if not os.path.exists('output'):
            os.makedirs('output')
        self.destination_file_name = 'output/result.xlsx'

    def initialize(self):
        self.ws1 = self.wb.active
        self.ws1.title = 'Agencies'
        self.ws1.append([
            'Agency name',
            'Spend Amounts'
        ])

    def save(self):
        self.wb.save(filename=self.destination_file_name)

    def write_to_file(self, dict_items, sheet):
        for key, value in dict_items:
            sheet.append([key, value])

    def read_from_file(self):
        pass

    def create_sheet(self, title):
        return self.wb.create_sheet(title=title)


class WebPageDataExtracter(SeleniumDriver):
    def __init__(self, agency_name):
        super().__init__()
        self.links_to_download_pdf = []
        self.agencies = {}
        self.url = 'https://itdashboard.gov/'
        self.dive_in_xpath = '//*[@id="node-23"]/div/div/div/div/div/div/div/a'
        self.agency_item_xpath = '//*[@id="agency-tiles-widget"]//a/img'
        self.agency_items_xpath = '//*[@id="agency-tiles-widget"]/div/div/div/div/div/div/div[1]/a'
        self.table_element_xpath = '//*[@id="investments-table-object_wrapper"]/div[3]/div[1]/div/table/thead/tr[2]/th[1]'
        self.select_element_xpath = '//*[@id="investments-table-object_length"]/label/select'
        self.all_options_xpath = '//*[@id="investments-table-object_length"]/label/select/option[4]'
        self.xlsx_handler = ExcelHandler()
        self.agency_name = agency_name
        self.agency_obj = None
        self.downloaded_pdf_file_names = []

    def parse_agencies(self):
        dive_in_button = self.get_element_if_exists_by_xpath(self.dive_in_xpath)
        dive_in_button.click()
        self.check_is_element_present_by_xpath(self.agency_item_xpath)
        agencies_elements = self.fetch_elements_if_exists_by_xpath(self.agency_items_xpath)
        for item in agencies_elements:
            name = item.find_element_by_xpath('./span[1]').text
            value = item.find_element_by_xpath('./span[2]').text
            self.agencies[name] = value
            if name == self.agency_name:
                self.agency_obj = item

    def parse_table(self):
        if self.is_loading(20) is None:
            table_body = self.get_element_if_exists_by_xpath('//*[@id="investments-table-object"]/tbody')
            for row in table_body.find_elements_by_xpath('./tr'):
                _row = []
                for index, col in enumerate(row.find_elements_by_xpath('./td')):
                    _row.append(col.text)
                    link_text = col.find_elements_by_link_text(col.text)
                    if link_text:
                        self.links_to_download_pdf.append(link_text[0].get_attribute('href'))
                self.ws2.append(_row)

    def check_agency(self):
        self.agency_obj.click()
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        self.check_is_element_present_by_xpath(self.table_element_xpath, 20)
        select_element = self.get_element_if_exists_by_xpath('//*[@id="investments-table-object_length"]/label/select')
        select_element.click()
        show_all_element = self.get_element_if_exists_by_xpath('//*[@id="investments-table-object_length"]/label/select/option[4]')
        show_all_element.click()
        self.parse_table()

    def download_pdf_files_from_links(self):
        for item in self.links_to_download_pdf:
            self.driver.get(item)
            self.check_is_element_present_by_xpath('//*[@id="business-case-pdf"]/a')
            element = self.get_element_if_exists_by_xpath('//*[@id="business-case-pdf"]/a')
            element.click()
            time.sleep(10)

    def pdf_reader(self):
        pass

    def execute(self):
        try:
            self.initialize_driver()
            self.xlsx_handler.initialize()
            self.ws2 = self.xlsx_handler.create_sheet('individual investments')
            self.driver.get(self.url)
            self.parse_agencies()
            self.xlsx_handler.write_to_file(self.agencies.items(), self.xlsx_handler.ws1)
            self.check_agency()
            self.download_pdf_files_from_links()
            self.xlsx_handler.save()
        finally:
            self.driver.quit()


if __name__ == '__main__':
    agency_name = sys.argv.pop()
    web_page_data_extracter = WebPageDataExtracter(agency_name=agency_name)
    web_page_data_extracter.execute()

